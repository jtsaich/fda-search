"""Tests for the ExcelService structured data extraction."""

import io
import csv
import pytest
import openpyxl

from services.excel_service import ExcelService


def _make_xlsx_bytes(sheets: dict[str, list[list]]) -> bytes:
    """Helper: create an in-memory .xlsx file from {sheet_name: [[row], ...]}."""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        if first:
            ws.title = name
            first = False
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv_bytes(rows: list[list]) -> bytes:
    """Helper: create in-memory CSV bytes from [[row], ...]."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


# ── XLSX tests ──


class TestExtractStructuredDataXlsx:
    def test_basic_extraction(self):
        data = _make_xlsx_bytes(
            {"Sales": [["Quarter", "Revenue", "Profit"], ["Q1", 100, 20], ["Q2", 200, 50]]}
        )
        result = ExcelService.extract_structured_data(data, "test.xlsx")

        assert result["filename"] == "test.xlsx"
        assert len(result["sheets"]) == 1
        sheet = result["sheets"][0]
        assert sheet["name"] == "Sales"
        assert sheet["headers"] == ["Quarter", "Revenue", "Profit"]
        assert len(sheet["data"]) == 2
        assert sheet["data"][0]["Quarter"] == "Q1"
        assert sheet["data"][0]["Revenue"] == 100
        assert sheet["data"][1]["Profit"] == 50

    def test_data_types_inferred(self):
        data = _make_xlsx_bytes(
            {"Data": [["Name", "Value", "Count"], ["A", "hello", 5], ["B", "world", 10]]}
        )
        result = ExcelService.extract_structured_data(data, "types.xlsx")
        types = result["sheets"][0]["data_types"]
        assert types["Name"] == "string"
        assert types["Value"] == "string"
        assert types["Count"] == "number"

    def test_empty_rows_skipped(self):
        data = _make_xlsx_bytes(
            {"Sheet1": [["Col1", "Col2"], [None, None], ["a", 1]]}
        )
        result = ExcelService.extract_structured_data(data, "sparse.xlsx")
        assert len(result["sheets"][0]["data"]) == 1
        assert result["sheets"][0]["data"][0]["Col1"] == "a"

    def test_multi_sheet(self):
        data = _make_xlsx_bytes(
            {
                "Sheet1": [["A", "B"], [1, 2]],
                "Sheet2": [["X", "Y"], [10, 20]],
            }
        )
        result = ExcelService.extract_structured_data(data, "multi.xlsx")
        assert len(result["sheets"]) == 2
        assert result["sheets"][0]["name"] == "Sheet1"
        assert result["sheets"][1]["name"] == "Sheet2"

    def test_empty_workbook(self):
        wb = openpyxl.Workbook()
        # Default sheet has no data, but has headers row - actually it's completely empty
        buf = io.BytesIO()
        wb.save(buf)
        result = ExcelService.extract_structured_data(buf.getvalue(), "empty.xlsx")
        assert result["sheets"] == []

    def test_none_headers_get_default_names(self):
        data = _make_xlsx_bytes({"Sheet1": [[None, "B", None], [1, 2, 3]]})
        result = ExcelService.extract_structured_data(data, "noheaders.xlsx")
        headers = result["sheets"][0]["headers"]
        assert headers[0] == "Column_0"
        assert headers[1] == "B"
        assert headers[2] == "Column_2"


# ── CSV tests ──


class TestExtractStructuredDataCsv:
    def test_basic_csv(self):
        data = _make_csv_bytes(
            [["Product", "Sales", "Region"], ["Widget", "100", "East"], ["Gadget", "200", "West"]]
        )
        result = ExcelService.extract_structured_data(data, "test.csv")

        assert result["filename"] == "test.csv"
        assert len(result["sheets"]) == 1
        sheet = result["sheets"][0]
        assert sheet["name"] == "Sheet1"
        assert sheet["headers"] == ["Product", "Sales", "Region"]
        assert len(sheet["data"]) == 2
        # Numbers should be coerced
        assert sheet["data"][0]["Sales"] == 100
        assert sheet["data"][0]["Product"] == "Widget"

    def test_csv_number_coercion(self):
        data = _make_csv_bytes([["Name", "Int", "Float"], ["A", "42", "3.14"]])
        result = ExcelService.extract_structured_data(data, "nums.csv")
        row = result["sheets"][0]["data"][0]
        assert row["Int"] == 42
        assert row["Float"] == 3.14
        assert result["sheets"][0]["data_types"]["Int"] == "number"
        assert result["sheets"][0]["data_types"]["Float"] == "number"
        assert result["sheets"][0]["data_types"]["Name"] == "string"

    def test_csv_empty_cells(self):
        data = _make_csv_bytes([["A", "B"], ["", "1"], ["x", ""]])
        result = ExcelService.extract_structured_data(data, "sparse.csv")
        assert len(result["sheets"][0]["data"]) == 2
        assert result["sheets"][0]["data"][0]["A"] is None
        assert result["sheets"][0]["data"][0]["B"] == 1

    def test_csv_empty_file(self):
        data = b""
        result = ExcelService.extract_structured_data(data, "empty.csv")
        assert result["sheets"] == []


# ── Summary tests ──


class TestGenerateDataSummary:
    def test_summary_contains_key_info(self):
        data = _make_xlsx_bytes(
            {"Sales": [["Quarter", "Revenue"], ["Q1", 100], ["Q2", 200]]}
        )
        structured = ExcelService.extract_structured_data(data, "sales.xlsx")
        summary = ExcelService.generate_data_summary(structured)

        assert "sales.xlsx" in summary
        assert "Sales" in summary
        assert "Quarter" in summary
        assert "Revenue" in summary
        assert "Row count: 2" in summary
        assert "Sample data:" in summary

    def test_csv_summary(self):
        data = _make_csv_bytes([["City", "Population"], ["NYC", "8000000"]])
        structured = ExcelService.extract_structured_data(data, "cities.csv")
        summary = ExcelService.generate_data_summary(structured)

        assert "cities.csv" in summary
        assert "City" in summary
        assert "Population" in summary
